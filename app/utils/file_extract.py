import logging
from pathlib import Path

logger = logging.getLogger(__name__)

ALLOWED_FILE_EXTENSIONS = {".xlsx", ".docx", ".txt", ".pdf"}


def extract_text_from_file(file_path: str) -> str:
    """Извлекает текстовое содержимое из файла по расширению."""
    path = Path(file_path)
    ext = path.suffix.lower()

    if ext == ".txt":
        return _extract_txt(path)
    if ext == ".pdf":
        return _extract_pdf(path)
    if ext == ".docx":
        return _extract_docx(path)
    if ext == ".xlsx":
        return _extract_xlsx(path)

    logger.warning("Unsupported file extension: %s", ext)
    return ""


def _extract_txt(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def _extract_pdf(path: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    parts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            parts.append(text)
    return "\n".join(parts)


def _extract_docx(path: Path) -> str:
    from docx import Document

    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def _extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                parts.append(" ".join(cells))
    wb.close()
    return "\n".join(parts)
