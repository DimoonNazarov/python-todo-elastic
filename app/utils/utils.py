import os
import hashlib
import random
import string
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from fastapi import UploadFile
from fastapi import HTTPException
from fastapi import status
from loguru import logger

from app.models import Todo
from app.schemas import TodoSource


def export_todos(todos: list[Todo], file_path: str = "data/todos.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("todos", 0)

    headers = [
        "title",
        "details",
        "completed",
        "tag",
        "created_at",
        "completed_at",
        "due_at",
        "updated_at",
        "updated_by",
        "source",
        "spacy_summary",
        "image_path",
        "image_hash",
    ]
    for index, header in enumerate(headers):
        ws.column_dimensions[f"{chr(index + 65)}"].width = len(header) + 5
    ws.append(headers)
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center")

    def fmt(dt):
        return dt.strftime("%Y-%m-%d %H:%M:%S") if dt is not None else ""

    for todo in todos:
        ws.append(
            [
                todo.title,
                todo.details,
                "Выполнено" if todo.completed else "Не выполнено",
                todo.tag,
                fmt(todo.created_at),
                fmt(todo.completed_at),
                fmt(todo.due_at),
                fmt(todo.updated_at),
                todo.updated_by,
                todo.source,
                todo.spacy_summary,
                todo.image_path,
                todo.image_hash,
            ]
        )

    # Скрываем служебные колонки
    for hidden_col in ("image_hash",):
        for cell in ws[1]:
            if cell.value == hidden_col:
                ws.column_dimensions[cell.column_letter].hidden = True

    wb.save(file_path)


def import_todos(file_path) -> list[Todo]:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    todos = []
    column_index = None

    for cell in sheet[1]:
        if cell.value == "image_hash":
            column_index = cell.column_letter
            break

    if column_index:
        sheet.column_dimensions[column_index].hidden = False

    for row in sheet.iter_rows(min_row=2, values_only=True):
        (
            title,
            details,
            completed,
            tag,
            created_at,
            completed_at,
            source,
            image_path,
            image_hash,
        ) = row

        if not completed and completed_at is not None:
            print(f"Ошибка: Задача с ID {id} не завершена, но дата выполнения указана.")
            continue

        created_at = created_at if isinstance(created_at, datetime) else None
        completed_at = completed_at if isinstance(completed_at, datetime) else None

        todo = Todo()
        todo.title = title
        todo.details = details if details is not None else ""
        todo.completed = True if completed == "Выполнено" else False
        todo.tag = tag
        todo.created_at = created_at
        todo.completed_at = completed_at
        todo.source = TodoSource.imported
        todo.image_path = image_path
        todo.image_hash = image_hash
        todos.append(todo)

    workbook.close()

    return todos


def hash_text(text: str) -> str:
    """Возвращает MD5-хеш строки текста."""
    return hashlib.md5(text.encode("utf-8")).hexdigest()


async def hash_image(image: UploadFile):
    try:
        img_bytes = await image.read()
        img_hash = hashlib.md5(img_bytes).hexdigest()
    except Exception as e:
        raise ValueError(f"Cannot identify image file: {e}")
    return img_hash


async def load_image(image: UploadFile, random_filename: str) -> None:
    """Load image"""
    file_location = os.path.join("./images/", random_filename)
    try:
        with open(file_location, "wb") as file:
            file.write(await image.read())
    except Exception as e:
        logger.error(f"Error saving image {random_filename}: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Image saving failed",
        )


def generate_random_filename(length: int = 10) -> str:
    """Generate a random filename of specified length."""
    characters = string.ascii_letters + string.digits
    return "".join(random.choice(characters) for _ in range(length))


async def delete_image(image_path: str) -> None:
    """Delete image"""
    try:
        full_path = os.path.join("./images/", image_path)
        if os.path.exists(full_path):
            os.remove(full_path)
            logger.info(f"Image deleted successfully: {image_path}")
        else:
            logger.warning(f"Image not found for deletion: {image_path}")
    except Exception as e:
        logger.error(f"Error deleting image {image_path}: {e}")


def create_dirs():
    if not os.path.exists("data"):
        os.mkdir("data")

    if not os.path.exists("images"):
        os.mkdir("images")
